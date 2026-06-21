# -*- coding: utf-8 -*-
"""
Xuat danh sach job da loc ra file Excel (.xlsx) de tu day len Google Sheets.
Nguon uu tien: data/jobs_filtered.json -> data/jobs_ranked.json
Chay: python export_excel.py
"""
import os, sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

SRC_CANDIDATES = ["data/jobs_filtered.json", "data/jobs_ranked.json"]
OUT = "data/Jobs_phu_hop.xlsx"

HEADERS = [
    ("STT", 6),
    ("Phu hop %", 11),
    ("Vi tri", 42),
    ("Cong ty", 26),
    ("Dia diem", 24),
    ("Luong", 16),
    ("Dang", 12),
    ("Link LinkedIn", 20),
    ("Nhan xet", 50),
    ("Diem khop", 45),
    ("Con thieu", 38),
    ("Keyword nen them", 28),
    ("Tieu chi (✔/✘)", 38),
]

BLUE = "0A66C2"
LIGHT = "EAF1FB"


def load():
    for p in SRC_CANDIDATES:
        if os.path.exists(p):
            return json.load(open(p, encoding="utf-8")), p
    raise SystemExit("Khong tim thay file ket qua. Chay match_jobs.py truoc.")


def main():
    jobs, src = load()
    wb = Workbook()
    ws = wb.active
    ws.title = "Job phu hop"

    thin = Side(style="thin", color="D0D7E2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    for c, (name, w) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = w

    # Rows
    for i, j in enumerate(jobs, 1):
        m = j.get("match", {})
        loc = j.get("work_location_detail") or j.get("location") or ""
        row = [
            i,
            m.get("match_score", ""),
            j.get("title", ""),
            j.get("company", ""),
            loc,
            j.get("salary") or "",
            j.get("posted") or "",
            None,  # link (set hyperlink below)
            m.get("one_line_reason", ""),
            "; ".join(m.get("strengths", [])),
            "; ".join(m.get("gaps", [])),
            ", ".join(m.get("keywords_missing", [])),
            "  ".join(m.get("criteria_flags", [])),
        ]
        r = i + 1
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(c >= 3))
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT)
        # cot Phu hop %: to mau theo diem
        sc = m.get("match_score", 0) or 0
        scell = ws.cell(row=r, column=2)
        scell.alignment = Alignment(horizontal="center", vertical="center")
        scell.font = Font(bold=True, color=("2E7D32" if sc >= 70 else "E65100" if sc >= 50 else "C62828"))
        # cot Link: hyperlink
        url = j.get("url", "")
        lcell = ws.cell(row=r, column=8, value="Mo job ↗" if url else "")
        if url:
            lcell.hyperlink = url
            lcell.font = Font(color=BLUE, underline="single")
        lcell.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze header + auto filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(jobs)+1}"
    ws.row_dimensions[1].height = 28

    os.makedirs("data", exist_ok=True)
    wb.save(OUT)
    print(f"[OK] Da xuat {len(jobs)} job -> {OUT}  (nguon: {src})")
    print(f"     Mo file roi: File > Import vao Google Sheets la xong.")


if __name__ == "__main__":
    main()
